import streamlit as st
import pdfplumber
import pandas as pd
import numpy as np
from io import BytesIO
import tempfile
import os
import re
from difflib import SequenceMatcher
import datetime
import locale
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
# from openpyxl.formatting.rule import DataBarRule # Removido DataBarRule
import openpyxl

# Configurar locale para português brasileiro
try:
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
except:
    try:
        locale.setlocale(locale.LC_ALL, 'Portuguese_Brazil.1252')
    except:
        pass  # Fallback para o locale padrão

def similar(a, b):
    """Calcula similaridade entre duas strings"""
    return SequenceMatcher(None, a, b).ratio()

def clean_columns(columns):
    """Garante que os nomes das colunas sejam únicos e limpos"""
    # Primeiro, limpar e normalizar os nomes das colunas
    cleaned_cols = []
    for col in columns:
        # Garantir que a coluna não seja None
        if col is None:
            col = "Coluna"
        else:
            col = str(col).strip()
            # Remover caracteres especiais e múltiplos espaços
            col = re.sub(r'\s+', ' ', col)
            if col == '':
                col = "Coluna"
        cleaned_cols.append(col)
    
    # Depois, garantir que sejam únicos
    seen = {}
    new_columns = []
    for col in cleaned_cols:
        new_col = col
        count = 1
        while new_col in seen:
            new_col = f"{col}_{count}"
            count += 1
        seen[new_col] = True
        new_columns.append(new_col)
    
    return new_columns

def detect_column_type(column_name, sample_values):
    """Detecta o tipo de dados de uma coluna com base no nome e valores"""
    column_name = str(column_name).lower()
    
    # Padrões para reconhecimento de tipos
    money_patterns = ['valor', 'preço', 'custo', 'total', 'r$', 'reais', 'saldo', 'montante']
    date_patterns = ['data', 'período', 'mês', 'ano', 'dia', 'dt_']
    percent_patterns = ['percentual', '%', 'porcentagem', 'taxa']
    quantity_patterns = ['quantidade', 'qtd', 'qtde', 'volume', 'número', 'num']
    
    # Verificar pelo nome da coluna
    if any(pattern in column_name for pattern in money_patterns):
        return 'money'
    elif any(pattern in column_name for pattern in date_patterns):
        return 'date'
    elif any(pattern in column_name for pattern in percent_patterns):
        return 'percent'
    elif any(pattern in column_name for pattern in quantity_patterns):
        return 'number'
    
    # Se não foi possível determinar pelo nome, tentar pelos valores
    numeric_count = 0
    date_count = 0
    percent_count = 0
    money_count = 0
    
    # Padrões regex para detecção
    money_regex = r'^R?\$?\s*\d+[.,]\d+$|^\d+[.,]\d+\s*R?\$?$'
    percent_regex = r'^\d+[.,]?\d*\s*\%$'
    date_regex = r'^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$|^\d{2,4}[/-]\d{1,2}[/-]\d{1,2}$'
    number_regex = r'^\d+[.,]?\d*$' # Corrigido: adicionado o fechamento da aspa simples
    
    for value in sample_values:
        if value is None or pd.isna(value) or str(value).strip() == '':
            continue
            
        value_str = str(value).strip()
        
        if re.match(money_regex, value_str):
            money_count += 1
        elif re.match(percent_regex, value_str):
            percent_count += 1
        elif re.match(date_regex, value_str):
            date_count += 1
        elif re.match(number_regex, value_str):
            numeric_count += 1
    
    # Determinar o tipo com base na contagem
    total_valid = money_count + percent_count + date_count + numeric_count
    if total_valid == 0:
        return 'text'
    
    threshold = 0.6  # 60% dos valores precisam corresponder ao padrão
    
    if money_count / total_valid >= threshold:
        return 'money'
    elif percent_count / total_valid >= threshold:
        return 'percent'
    elif date_count / total_valid >= threshold:
        return 'date'
    elif numeric_count / total_valid >= threshold:
        return 'number'
    
    return 'text'

def convert_to_numeric(value):
    """Converte um valor para numérico, removendo caracteres não numéricos"""
    if value is None or pd.isna(value):
        return None
        
    value_str = str(value).strip()
    
    # Remover símbolos de moeda e outros caracteres não numéricos
    value_str = re.sub(r'[^\d,.-]', '', value_str)
    
    # Substituir vírgula por ponto para conversão
    value_str = value_str.replace(',', '.')
    
    try:
        return float(value_str)
    except:
        return None

def convert_to_date(value):
    """Tenta converter um valor para data"""
    if value is None or pd.isna(value):
        return None
        
    value_str = str(value).strip()
    
    # Formatos comuns de data no Brasil
    date_formats = [
        '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d', '%Y-%m-%d',
        '%d/%m/%y', '%d-%m-%y', '%y/%m/%d', '%y-%m-%d'
    ]
    
    for fmt in date_formats:
        try:
            return datetime.datetime.strptime(value_str, fmt).date()
        except:
            continue
            
    return None

def convert_to_percent(value):
    """Converte um valor para percentual"""
    if value is None or pd.isna(value):
        return None
        
    value_str = str(value).strip()
    
    # Remover o símbolo de percentual
    value_str = value_str.replace('%', '')
    
    # Substituir vírgula por ponto para conversão
    value_str = value_str.replace(',', '.')
    
    try:
        return float(value_str) / 100.0
    except:
        return None

def process_dataframe(df):
    """Processa o DataFrame para converter tipos de dados"""
    # Criar uma cópia para não modificar o original
    processed_df = df.copy()
    
    # Detectar e converter tipos de dados para cada coluna
    for col in processed_df.columns:
        # Pular colunas de metadados
        if col in ['Origem', 'Página', 'Tabela']:
            continue
            
        # Obter amostra de valores não nulos
        sample_values = processed_df[col].dropna().astype(str).tolist()[:20]
        if not sample_values:
            continue
            
        # Detectar tipo da coluna
        col_type = detect_column_type(col, sample_values)
        
        # Converter valores com base no tipo detectado
        if col_type == 'money':
            processed_df[col] = processed_df[col].apply(convert_to_numeric)
        elif col_type == 'number':
            processed_df[col] = processed_df[col].apply(convert_to_numeric)
        elif col_type == 'date':
            processed_df[col] = processed_df[col].apply(convert_to_date)
        elif col_type == 'percent':
            processed_df[col] = processed_df[col].apply(convert_to_percent)
    
    return processed_df

def detect_table_type(df):
    """Detecta o tipo de tabela com base nas colunas e conteúdo"""
    # Implementação básica - pode ser expandida com mais heurísticas
    cols = [str(col).lower() for col in df.columns]
    
    # Verificar padrões comuns em colunas
    financial_keywords = ['valor', 'preço', 'custo', 'total', 'r$', 'reais', 'saldo']
    date_keywords = ['data', 'período', 'mês', 'ano', 'dia']
    person_keywords = ['nome', 'pessoa', 'cliente', 'funcionário', 'cpf', 'cnpj']
    
    financial_score = sum(1 for kw in financial_keywords if any(kw in col for col in cols))
    date_score = sum(1 for kw in date_keywords if any(kw in col for col in cols))
    person_score = sum(1 for kw in person_keywords if any(kw in col for col in cols))
    
    if financial_score > max(date_score, person_score):
        return "financial"
    elif date_score > max(financial_score, person_score):
        return "date"
    elif person_score > max(financial_score, date_score):
        return "person"
    else:
        return "general"

def standardize_columns(all_tables):
    """Padroniza as colunas entre tabelas similares para facilitar a concatenação"""
    if not all_tables:
        return []
    
    # Agrupar tabelas por tipo
    table_groups = {}
    for df in all_tables:
        table_type = detect_table_type(df)
        if table_type not in table_groups:
            table_groups[table_type] = []
        table_groups[table_type].append(df)
    
    standardized_tables = []
    
    # Processar cada grupo separadamente
    for table_type, tables in table_groups.items():
        # Encontrar colunas comuns no grupo
        column_freq = {}
        for df in tables:
            for col in df.columns:
                col_lower = str(col).lower()
                if col_lower not in column_freq:
                    column_freq[col_lower] = {'count': 0, 'variations': {}}
                column_freq[col_lower]['count'] += 1
                if col not in column_freq[col_lower]['variations']:
                    column_freq[col_lower]['variations'][col] = 0
                column_freq[col_lower]['variations'][col] += 1
        
        # Determinar o nome padrão para cada coluna (o mais frequente)
        standard_names = {}
        for col_lower, data in column_freq.items():
            variations = data['variations']
            standard_names[col_lower] = max(variations.items(), key=lambda x: x[1])[0]
        
        # Padronizar cada tabela no grupo
        for df in tables:
            # Criar mapeamento de nomes de colunas
            col_mapping = {}
            for col in df.columns:
                col_lower = str(col).lower()
                if col_lower in standard_names:
                    col_mapping[col] = standard_names[col_lower]
            
            # Renomear colunas
            if col_mapping:
                df = df.rename(columns=col_mapping)
            
            standardized_tables.append(df)
    
    return standardized_tables

def extract_tables_with_context(file):
    """Extrai tabelas com contexto do texto anterior"""
    all_tables_data = []  # Lista para armazenar DataFrames com contexto
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
        tmp_file.write(file.getvalue())
        tmp_path = tmp_file.name
    
    try:
        with pdfplumber.open(tmp_path) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                # Tentar diferentes estratégias de extração
                extraction_strategies = [
                    {"vertical_strategy": "text", "horizontal_strategy": "text"},
                    {"vertical_strategy": "lines", "horizontal_strategy": "lines"},
                    {"vertical_strategy": "lines_strict", "horizontal_strategy": "lines_strict"}
                ]
                
                tables = []
                for strategy in extraction_strategies:
                    try:
                        tables = page.extract_tables(strategy)
                        if tables and any(len(t) > 1 for t in tables):
                            break  # Encontrou tabelas válidas, usar esta estratégia
                    except Exception:
                        continue  # Tentar próxima estratégia
                
                page_text = page.extract_text()
                
                for table_num, table in enumerate(tables, 1):
                    if not table or len(table) <= 1:  # Ignorar tabelas vazias
                        continue
                        
                    try:
                        # Processar tabela
                        headers = table[0]
                        data = table[1:]
                        
                        # Garantir headers únicos e limpos
                        headers = clean_columns(headers)
                        
                        # Criar DataFrame
                        df = pd.DataFrame(data, columns=headers)
                        
                        # Limpar dados
                        # Remover linhas vazias ou com muitos valores nulos
                        df = df.dropna(how='all').reset_index(drop=True)
                        df = df.loc[df.apply(lambda x: x.astype(str).str.strip().ne('').sum() > len(x) * 0.3, axis=1)]
                        
                        if not df.empty:
                            # Obter contexto (texto antes da tabela)
                            context = f"Página {page_num}, Tabela {table_num}"
                            try:
                                # Estimar a posição Y da tabela
                                y_position = 0
                                for word in page.extract_words():
                                    if any(cell and str(cell).strip() in word['text'] for row in table for cell in row):
                                        y_position = word['top']
                                        break
                                
                                if y_position > 0:
                                    upper_part = page.crop((0, 0, page.width, y_position))
                                    context_text = upper_part.extract_text()
                                    if context_text:
                                        context_lines = [line.strip() for line in context_text.split('\n') if line.strip()]
                                        if context_lines:
                                            # Pegar as últimas linhas como contexto (ajustado para pegar mais texto)
                                            context = ' '.join(context_lines[-7:]) # Aumentado para 7 linhas
                                        else:
                                            # Fallback se não encontrar texto antes da tabela mas houver texto na página
                                            if page_text.strip():
                                                context = ' '.join(page_text.strip().split('\n')[-3:])
                                    
                            except Exception:
                                # Se falhar, manter o contexto padrão
                                pass
                            
                            # Adicionar colunas de metadados
                            df['Origem'] = context
                            df['Página'] = page_num
                            df['Tabela'] = table_num
                            
                            all_tables_data.append(df)
                    except Exception as e:
                        st.warning(f"Ignorando tabela na página {page_num} devido a erro: {str(e)}")
                        continue
    finally:
        try:
            os.unlink(tmp_path)
        except:
            pass
    
    # Não agrupar por similaridade, apenas retornar a lista de tabelas
    return all_tables_data

def format_excel_worksheet(worksheet, df):
    """Aplica formatação a uma planilha Excel (sem barras de dados)"""
    # Definir estilos
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Aplicar estilos ao cabeçalho
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Ajustar largura das colunas
    for i, column in enumerate(df.columns, 1):
        column_letter = get_column_letter(i)
        # Calcular largura baseada no conteúdo
        max_length = max(
            len(str(column)),
            df[column].astype(str).apply(len).max() if not df.empty else 0
        )
        adjusted_width = min(max(max_length + 2, 10), 50)  # Entre 10 e 50 caracteres
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Aplicar formatação condicional para colunas numéricas
    for i, column in enumerate(df.columns, 1):
        column_letter = get_column_letter(i)
        
        # Detectar tipo de coluna
        # Pular colunas de metadados
        if column in ['Origem', 'Página', 'Tabela']:
             for cell in worksheet[column_letter][1:]:
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.border = border
             continue
            
        sample_values = df[column].dropna().astype(str).tolist()[:20]
        if not sample_values:
            # Aplicar bordas e alinhamento padrão mesmo se não houver dados
            for cell in worksheet[column_letter][1:]:
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.border = border
            continue
            
        col_type = detect_column_type(column, sample_values)
        
        # Aplicar formatação baseada no tipo
        if col_type == 'money':
            # Formato de moeda
            for cell in worksheet[column_letter][1:]:
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = 'R$ #,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                cell.border = border
            
        elif col_type == 'percent':
            # Formato percentual
            for cell in worksheet[column_letter][1:]:
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00%'
                    cell.alignment = Alignment(horizontal='right')
                cell.border = border
        elif col_type == 'number':
            # Formato numérico
            for cell in worksheet[column_letter][1:]:
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                cell.border = border
        elif col_type == 'date':
            # Formato de data
            for cell in worksheet[column_letter][1:]:
                if cell.value is not None and isinstance(cell.value, datetime.date):
                    cell.number_format = 'dd/mm/yyyy'
                    cell.alignment = Alignment(horizontal='center')
                cell.border = border
        else:
            # Texto
            for cell in worksheet[column_letter][1:]:
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.border = border
    
    # Congelar painéis para facilitar a navegação
    worksheet.freeze_panes = 'D2'  # Congelar cabeçalho e colunas de metadados
    
    # Adicionar filtros ao cabeçalho
    worksheet.auto_filter.ref = worksheet.dimensions

def create_excel_file(all_tables):
    """Cria arquivo Excel com todas as tabelas concatenadas em uma única planilha"""
    if not all_tables:
        # Criar planilha vazia como fallback
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            pd.DataFrame(['Nenhuma tabela válida foi encontrada']).to_excel(
                writer, sheet_name="Info", index=False)
        output.seek(0)
        return output
    
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Concatenar todas as tabelas em uma única planilha
        try:
            # Processar tipos de dados para cada tabela individualmente antes de concatenar
            processed_tables = [process_dataframe(df) for df in all_tables]
            
            # Coletar todas as colunas possíveis de todas as tabelas processadas
            all_columns = set()
            for df in processed_tables:
                all_columns.update(df.columns)
            
            # Garantir que todas as tabelas tenham as mesmas colunas antes de concatenar
            final_tables = []
            for df in processed_tables:
                # Criar cópia para não modificar o original
                std_df = df.copy()
                # Adicionar colunas faltantes com valores NaN
                for col in all_columns:
                    if col not in std_df.columns:
                        std_df[col] = pd.NA
                final_tables.append(std_df)
                
            # Concatenar todas as tabelas padronizadas
            combined_df = pd.concat(final_tables, ignore_index=True)
            
            # Reordenar colunas para que metadados fiquem primeiro
            metadata_cols = ['Origem', 'Página', 'Tabela']
            other_cols = [col for col in combined_df.columns if col not in metadata_cols]
            # Garantir que todas as colunas de metadados existam antes de reordenar
            ordered_cols = [col for col in metadata_cols if col in combined_df.columns] + other_cols
            combined_df = combined_df[ordered_cols]
            
            # Salvar na planilha principal
            combined_df.to_excel(writer, sheet_name="Todas as Tabelas", index=False)
            
            # Aplicar formatação à planilha principal
            workbook = writer.book
            worksheet = writer.sheets["Todas as Tabelas"]
            format_excel_worksheet(worksheet, combined_df)
            
        except Exception as e:
            # Fallback: criar planilha simples se a padronização ou concatenação falhar
            st.warning(f"Usando método alternativo de agrupamento devido a: {str(e)}")
            
            # Simplesmente concatenar sem padronização e processamento de tipos
            simple_combined = pd.concat(all_tables, ignore_index=True, sort=False)
            simple_combined.to_excel(writer, sheet_name="Todas as Tabelas", index=False)
            
            # Aplicar formatação básica
            worksheet = writer.sheets["Todas as Tabelas"]
            format_excel_worksheet(worksheet, simple_combined)
    
    output.seek(0)
    return output

def main():
    st.set_page_config(
        page_title="PDF para Excel Avançado",
        page_icon="📊",
        layout="centered"
    )
    
    st.title("📊 PDF para Excel Avançado")
    st.markdown("""
    **Conversor inteligente que:**
    - Extrai todas as tabelas do PDF
    - Identifica o contexto/título de cada tabela
    - Combina todas as tabelas em uma única planilha sequencialmente
    - Detecta e formata automaticamente tipos de dados (números, datas, moeda)
    - Aplica formatação profissional às planilhas (sem barras de dados)
    - Mantém a origem de cada linha de dados
    """)
    
    uploaded_file = st.file_uploader(
        "Carregue seu arquivo PDF",
        type="pdf",
        accept_multiple_files=False
    )
    
    if uploaded_file is not None:
        with st.spinner("Processando PDF..."):
            try:
                # A função extract_tables_with_context agora retorna apenas a lista de tabelas
                all_tables = extract_tables_with_context(uploaded_file)
                
                if not all_tables:
                    st.warning("⚠️ Nenhuma tabela encontrada no PDF.")
                else:
                    total_tables = len(all_tables)
                    st.success(f"✅ {total_tables} tabelas processadas!")
                    
                    # A função create_excel_file agora recebe apenas a lista de tabelas
                    output = create_excel_file(all_tables)
                    file_name = uploaded_file.name.replace('.pdf', '') + '_tabelas_consolidadas.xlsx'
                    
                    st.download_button(
                        label="⬇️ Baixar Arquivo Excel Consolidado",
                        data=output,
                        file_name=file_name,
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    
                    st.subheader("📋 Pré-visualização da Tabela Consolidada")
                    try:
                        # Criar uma prévia da tabela combinada
                        preview_combined = pd.concat(all_tables, ignore_index=True).head(20)
                        st.dataframe(preview_combined)
                    except Exception as e:
                        st.warning(f"Não foi possível gerar pré-visualização: {str(e)}")
                        
            except Exception as e:
                st.error(f"❌ Falha no processamento: {str(e)}")

if __name__ == "__main__":
    main()

